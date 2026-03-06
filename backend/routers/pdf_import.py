"""
Router for PDF-based bulk data import.

Allows Super Admins to upload PDF files (departments, faculty, subjects,
rooms, venues, time-slots, batches) and extract + insert data into the DB.
"""
import io
import uuid
import tempfile
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from dependencies import get_db, require_super_admin
from models.user import User, UserRole
from models.college import College, Department
from models.faculty import Faculty
from models.subject import Subject
from models.room import Room, RoomType, Venue, VenueType
from models.timeslot import TimeSlotConfig, SlotType
from models.batch import Batch
from utils.security import hash_password

try:
    from scripts.pdf_parser import (
        parse_departments,
        parse_faculty,
        parse_subjects,
        parse_rooms,
        parse_venues,
        parse_timeslots,
        parse_batches,
        _map_header,
        _clean_str,
        _int_or,
        _yn_to_bool,
        _parse_list,
    )
except ImportError:
    import sys, os
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "scripts"))
    from pdf_parser import (
        parse_departments,
        parse_faculty,
        parse_subjects,
        parse_rooms,
        parse_venues,
        parse_timeslots,
        parse_batches,
        _map_header,
        _clean_str,
        _int_or,
        _yn_to_bool,
        _parse_list,
    )


router = APIRouter(prefix="/pdf-import", tags=["PDF Import"])

ROOM_TYPE_MAP = {
    "CLASSROOM": RoomType.CLASSROOM,
    "LAB": RoomType.LAB,
    "SEMINAR": RoomType.SEMINAR,
}

VENUE_TYPE_MAP = {
    "HALL": VenueType.HALL,
    "LAB": VenueType.LAB,
    "OUTDOOR": VenueType.OUTDOOR,
}

SLOT_TYPE_MAP = {
    "LECTURE": SlotType.LECTURE,
    "LAB": SlotType.LAB,
    "BREAK": SlotType.BREAK,
}

ALLOWED_CATEGORIES = [
    "departments", "faculty", "subjects",
    "rooms", "venues", "timeslots", "batches",
]


def _id() -> str:
    return str(uuid.uuid4())


async def _save_upload_to_temp(upload: UploadFile) -> Path:
    """Save uploaded file to a temp path and return it."""
    suffix = Path(upload.filename or "upload.pdf").suffix
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    content = await upload.read()
    tmp.write(content)
    tmp.close()
    return Path(tmp.name)


async def _get_dept_map(db: AsyncSession, college_id: str) -> dict[str, str]:
    """Return {dept_code_upper: dept_id} for the college."""
    result = await db.execute(
        select(Department).where(Department.college_id == college_id)
    )
    return {d.code.upper(): d.dept_id for d in result.scalars().all()}


# ── Sheet name → category mapping for Excel import ──────────

SHEET_CATEGORY_MAP: dict[str, str] = {
    "departments": "departments",
    "faculty": "faculty",
    "subjects": "subjects",
    "rooms": "rooms",
    "venues": "venues",
    "time slots": "timeslots",
    "timeslots": "timeslots",
    "batches": "batches",
}


def _convert_raw_rows(category: str, raw_rows: list[dict[str, str]]) -> list[dict]:
    """Convert raw {canonical_header: str_value} rows to typed dicts matching parse_* output."""
    if category == "departments":
        return [{"code": r.get("code", ""), "name": r.get("name", "")} for r in raw_rows]
    elif category == "faculty":
        return [{
            "department": _clean_str(r.get("department")),
            "name": _clean_str(r.get("name")),
            "employee_id": _clean_str(r.get("employeeid")),
            "email": _clean_str(r.get("email")),
            "phone": _clean_str(r.get("phone")),
            "expertise": _parse_list(r.get("expertise")),
            "max_load": _int_or(r.get("maxload"), 24),
            "preferred_time": _clean_str(r.get("preferredtime"), "any"),
            "role": _clean_str(r.get("role"), "faculty").lower(),
        } for r in raw_rows]
    elif category == "subjects":
        return [{
            "department": _clean_str(r.get("department")),
            "name": _clean_str(r.get("name")),
            "code": _clean_str(r.get("code")),
            "semester": _int_or(r.get("semester")),
            "credits": _int_or(r.get("credits")),
            "lecture_hours": _int_or(r.get("lecturehours")),
            "lab_hours": _int_or(r.get("labhours")),
            "batch_size": _int_or(r.get("batchsize"), 60),
        } for r in raw_rows]
    elif category == "rooms":
        return [{
            "name": _clean_str(r.get("name")),
            "capacity": _int_or(r.get("capacity"), 60),
            "type": _clean_str(r.get("type"), "CLASSROOM").upper(),
            "projector": _yn_to_bool(r.get("projector")),
            "computers": _yn_to_bool(r.get("computers")),
            "ac": _yn_to_bool(r.get("ac")),
        } for r in raw_rows]
    elif category == "venues":
        return [{
            "name": _clean_str(r.get("name")),
            "capacity": _int_or(r.get("capacity"), 100),
            "type": _clean_str(r.get("type"), "HALL").upper(),
        } for r in raw_rows]
    elif category == "timeslots":
        return [{
            "order": _int_or(r.get("order")),
            "label": _clean_str(r.get("label")),
            "start_time": _clean_str(r.get("starttime")),
            "end_time": _clean_str(r.get("endtime")),
            "type": _clean_str(r.get("type"), "LECTURE").upper(),
        } for r in raw_rows]
    elif category == "batches":
        return [{
            "department": _clean_str(r.get("department")),
            "semester": _int_or(r.get("semester")),
            "name": _clean_str(r.get("name")),
            "size": _int_or(r.get("size"), 20),
        } for r in raw_rows]
    return []


def _parse_excel_file(file_path: Path) -> dict[str, list[dict]]:
    """Parse all sheets from an Excel file, returning {category: typed_rows}."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    results: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        category = SHEET_CATEGORY_MAP.get(sheet_name.strip().lower())
        if not category:
            continue

        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            continue
        headers = [_map_header(str(h)) if h else "" for h in header_row]

        raw_rows: list[dict[str, str]] = []
        for row in rows_iter:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            row_dict: dict[str, str] = {}
            for i, hdr in enumerate(headers):
                if hdr and i < len(row):
                    row_dict[hdr] = _clean_str(row[i])
            if row_dict:
                raw_rows.append(row_dict)

        if raw_rows:
            results[category] = _convert_raw_rows(category, raw_rows)

    wb.close()
    return results


async def _run_import(db: AsyncSession, category: str, rows: list[dict], college_id: str) -> dict:
    """Dispatch to the correct import handler."""
    handlers = {
        "departments": _import_departments,
        "faculty": _import_faculty,
        "subjects": _import_subjects,
        "rooms": _import_rooms,
        "venues": _import_venues,
        "timeslots": _import_timeslots,
        "batches": _import_batches,
    }
    handler = handlers.get(category)
    if not handler:
        raise HTTPException(status_code=400, detail=f"Unknown category: {category}")
    return await handler(db, rows, college_id)


@router.post("/upload")
async def upload_pdf(
    file: UploadFile = File(...),
    category: str = Form(...),
    current_user: User = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Upload a PDF file and import data for the given category.

    **category** must be one of:
    departments, faculty, subjects, rooms, venues, timeslots, batches
    """
    category = category.strip().lower()
    if category not in ALLOWED_CATEGORIES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid category '{category}'. Must be one of: {', '.join(ALLOWED_CATEGORIES)}",
        )

    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted")

    # Save to temp file for pdfplumber
    tmp_path = await _save_upload_to_temp(file)
    college_id = current_user.college_id

    try:
        parser_map = {
            "departments": parse_departments,
            "faculty": parse_faculty,
            "subjects": parse_subjects,
            "rooms": parse_rooms,
            "venues": parse_venues,
            "timeslots": parse_timeslots,
            "batches": parse_batches,
        }
        rows = parser_map[category](str(tmp_path))
        if not rows:
            raise HTTPException(status_code=422, detail=f"No {category} data found in PDF")
        return await _run_import(db, category, rows, college_id)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Failed to parse PDF: {str(e)}")
    finally:
        tmp_path.unlink(missing_ok=True)


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Upload a single Excel file with multiple sheets to import all data at once.
    Sheet names should match: Departments, Faculty, Subjects, Rooms, Venues,
    Time Slots, Batches.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx) are accepted")

    tmp_path = await _save_upload_to_temp(file)
    college_id = current_user.college_id

    try:
        parsed = _parse_excel_file(tmp_path)
        if not parsed:
            raise HTTPException(
                status_code=422,
                detail="No valid sheets found. Expected sheets: Departments, Faculty, Subjects, Rooms, Venues, Time Slots, Batches",
            )

        import_order = ["departments", "faculty", "subjects", "rooms", "venues", "timeslots", "batches"]
        results = []
        for category in import_order:
            rows = parsed.get(category)
            if not rows:
                continue
            result = await _run_import(db, category, rows, college_id)
            results.append(result)

        return {"results": results, "sheets_processed": len(results)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Failed to parse Excel file: {str(e)}")
    finally:
        tmp_path.unlink(missing_ok=True)


@router.get("/categories")
async def list_categories(
    current_user: User = Depends(require_super_admin),
):
    """Return allowed import categories with descriptions."""
    return [
        {"key": "departments", "label": "Departments", "columns": "Code, Name"},
        {"key": "faculty", "label": "Faculty", "columns": "Department, Name, EmployeeID, Email, Phone, Expertise, MaxLoad, PreferredTime, Role"},
        {"key": "subjects", "label": "Subjects", "columns": "Department, Name, Code, Semester, Credits, LectureHours, LabHours, BatchSize"},
        {"key": "rooms", "label": "Rooms & Labs", "columns": "Name, Capacity, Type (CLASSROOM/LAB/SEMINAR), Projector (Y/N), Computers (Y/N), AC (Y/N)"},
        {"key": "venues", "label": "Exam Venues", "columns": "Name, Capacity, Type (HALL/LAB/OUTDOOR)"},
        {"key": "timeslots", "label": "Time Slots", "columns": "Order, Label, StartTime, EndTime, Type (LECTURE/LAB/BREAK)"},
        {"key": "batches", "label": "Batches", "columns": "Department, Semester, Name, Size"},
    ]


# ── Template definitions for downloadable files ────────────────

TEMPLATE_SHEETS: dict[str, dict] = {
    "Departments": {
        "headers": ["Code", "Name"],
        "sample": [["CP", "Computer Engineering"], ["EC", "Electronics & Communication"]],
        "widths": [15, 45],
    },
    "Faculty": {
        "headers": ["Department", "Name", "EmployeeID", "Email", "Phone", "Expertise", "MaxLoad", "PreferredTime", "Role"],
        "sample": [["CP", "Dr. Example Faculty", "CP001", "faculty@college.edu.in", "+919876500001", "AI, ML", "24", "morning", "Faculty"]],
        "widths": [14, 25, 14, 30, 18, 18, 10, 16, 10],
    },
    "Subjects": {
        "headers": ["Department", "Name", "Code", "Semester", "Credits", "LectureHours", "LabHours", "BatchSize"],
        "sample": [["CP", "Data Structures", "01030100", "3", "4", "3", "2", "60"]],
        "widths": [14, 30, 14, 12, 10, 16, 12, 12],
    },
    "Rooms": {
        "headers": ["Name", "Capacity", "Type", "Projector", "Computers", "AC"],
        "sample": [["Room 101", "60", "CLASSROOM", "Y", "N", "Y"]],
        "widths": [22, 12, 16, 12, 14, 8],
    },
    "Venues": {
        "headers": ["Name", "Capacity", "Type"],
        "sample": [["Main Exam Hall", "250", "HALL"]],
        "widths": [30, 12, 14],
    },
    "Time Slots": {
        "headers": ["Order", "Label", "StartTime", "EndTime", "Type"],
        "sample": [["1", "Period 1", "09:00", "10:00", "LECTURE"]],
        "widths": [10, 16, 14, 14, 14],
    },
    "Batches": {
        "headers": ["Department", "Semester", "Name", "Size"],
        "sample": [["CP", "1", "A", "20"]],
        "widths": [14, 12, 10, 10],
    },
}

EMPTY_ROWS_PER_SHEET = 20


@router.get("/template/pdf")
async def download_pdf_template(
    current_user: User = Depends(require_super_admin),
):
    """Download a PDF file with all template tables (headers + empty rows + 1 sample row)."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    for sheet_name, sheet in TEMPLATE_SHEETS.items():
        pdf.add_page("L")
        headers = sheet["headers"]
        sample = sheet["sample"]
        widths = sheet["widths"]

        # Title
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, sheet_name, new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(3)

        # Instruction
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 5, f"Fill in your data below. Row 1 is a sample — replace or delete it.", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

        n = len(headers)
        row_h = 8

        # Header row (blue background)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(41, 98, 255)
        pdf.set_text_color(255, 255, 255)
        for i, h in enumerate(headers):
            pdf.cell(widths[i], row_h, h, border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

        # Sample row (light yellow)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_fill_color(255, 255, 230)
        for row in sample:
            for i, cell in enumerate(row):
                pdf.cell(widths[i], row_h, str(cell), border=1, fill=True)
            pdf.ln()

        # Empty rows
        pdf.set_fill_color(255, 255, 255)
        for _ in range(EMPTY_ROWS_PER_SHEET):
            for i in range(n):
                pdf.cell(widths[i], row_h, "", border=1)
            pdf.ln()

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=import_template.pdf"},
    )


@router.get("/template/excel")
async def download_excel_template(
    current_user: User = Depends(require_super_admin),
):
    """Download an Excel file with all template sheets (headers + 1 sample row + empty rows)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2962FF", end_color="2962FF", fill_type="solid")
    sample_fill = PatternFill(start_color="FFFFE6", end_color="FFFFE6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    for sheet_name, sheet in TEMPLATE_SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        headers = sheet["headers"]
        sample = sheet["sample"]
        col_widths = sheet["widths"]

        # Set column widths
        for i, w in enumerate(col_widths):
            col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
            ws.column_dimensions[col_letter].width = w + 4

        # Header row
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Sample data row(s)
        for row_idx, row_data in enumerate(sample, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = sample_fill
                cell.border = thin_border

        # Empty rows with borders (for easy editing)
        start_row = 2 + len(sample)
        for row_idx in range(start_row, start_row + EMPTY_ROWS_PER_SHEET):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx, value="")
                cell.border = thin_border

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add data validation hints as a note on cell A1
        ws["A1"].comment = None  # clear any default

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_template.xlsx"},
    )


# ── Import handlers ────────────────────────────────────────────


async def _import_departments(db: AsyncSession, rows: list[dict], college_id: str):
    if not rows:
        return {"category": "departments", "inserted": 0, "skipped": 0, "total": 0}

    existing = await _get_dept_map(db, college_id)
    inserted, skipped = 0, 0

    for d in rows:
        code = d["code"].strip().upper()
        if not code:
            skipped += 1
            continue
        if code in existing:
            skipped += 1
            continue
        dept_id = _id()
        db.add(Department(
            dept_id=dept_id, college_id=college_id,
            name=d["name"].strip(), code=code,
        ))
        existing[code] = dept_id
        inserted += 1

    await db.commit()
    return {"category": "departments", "inserted": inserted, "skipped": skipped, "total": len(rows)}


async def _import_faculty(db: AsyncSession, rows: list[dict], college_id: str):
    if not rows:
        return {"category": "faculty", "inserted": 0, "skipped": 0, "total": 0}

    dept_map = await _get_dept_map(db, college_id)

    # Check existing emails to avoid duplicates
    result = await db.execute(
        select(User.email).where(User.college_id == college_id)
    )
    existing_emails = {e.lower() for e in result.scalars().all()}

    inserted, skipped = 0, 0
    warnings = []

    for f in rows:
        dept_code = f["department"].strip().upper()
        if dept_code not in dept_map:
            warnings.append(f"Skipped {f['name']}: department '{dept_code}' not found")
            skipped += 1
            continue

        email = f["email"].strip().lower()
        if not email:
            warnings.append(f"Skipped {f['name']}: no email")
            skipped += 1
            continue
        if email in existing_emails:
            skipped += 1
            continue

        dept_id = dept_map[dept_code]
        is_hod = f["role"] in ("hod", "dept_admin", "head")
        uid, fid = _id(), _id()

        db.add(User(
            user_id=uid, email=email, phone=f["phone"],
            hashed_password=hash_password("hod123" if is_hod else "faculty123"),
            full_name=f["name"],
            role=UserRole.DEPT_ADMIN if is_hod else UserRole.FACULTY,
            college_id=college_id, dept_id=dept_id,
        ))
        db.add(Faculty(
            faculty_id=fid, dept_id=dept_id, user_id=uid,
            name=f["name"], employee_id=f["employee_id"],
            expertise=f["expertise"],
            max_weekly_load=f["max_load"],
            preferred_time=f["preferred_time"],
        ))
        existing_emails.add(email)
        inserted += 1

    await db.commit()
    return {"category": "faculty", "inserted": inserted, "skipped": skipped, "total": len(rows), "warnings": warnings}


async def _import_subjects(db: AsyncSession, rows: list[dict], college_id: str):
    if not rows:
        return {"category": "subjects", "inserted": 0, "skipped": 0, "total": 0}

    dept_map = await _get_dept_map(db, college_id)

    # Check existing subject codes per department
    result = await db.execute(
        select(Subject.subject_code, Subject.dept_id).join(
            Department, Subject.dept_id == Department.dept_id
        ).where(Department.college_id == college_id)
    )
    existing = {(row.subject_code, row.dept_id) for row in result.all()}

    inserted, skipped = 0, 0
    warnings = []

    for s in rows:
        dept_code = s["department"].strip().upper()
        lh = s["lecture_hours"]
        lab_h = s["lab_hours"]

        # "Common" subjects go to all departments
        if dept_code in ("common", "all", "COMMON", "ALL", ""):
            target_depts = list(dept_map.items())
        elif dept_code in dept_map:
            target_depts = [(dept_code, dept_map[dept_code])]
        else:
            warnings.append(f"Skipped {s['name']}: department '{dept_code}' not found")
            skipped += 1
            continue

        for _, dept_id in target_depts:
            if (s["code"], dept_id) in existing:
                skipped += 1
                continue
            db.add(Subject(
                subject_id=_id(), dept_id=dept_id, name=s["name"],
                subject_code=s["code"], semester=s["semester"],
                credits=s["credits"],
                weekly_periods=lh + lab_h,
                lecture_hours=lh, lab_hours=lab_h,
                needs_lab=lab_h > 0, batch_size=s["batch_size"],
            ))
            existing.add((s["code"], dept_id))
            inserted += 1

    await db.commit()
    return {"category": "subjects", "inserted": inserted, "skipped": skipped, "total": len(rows), "warnings": warnings}


async def _import_rooms(db: AsyncSession, rows: list[dict], college_id: str):
    if not rows:
        return {"category": "rooms", "inserted": 0, "skipped": 0, "total": 0}

    # Check existing room names
    result = await db.execute(
        select(Room.name).where(Room.college_id == college_id)
    )
    existing_names = {n.lower() for n in result.scalars().all()}

    inserted, skipped = 0, 0

    for r in rows:
        name = r["name"].strip()
        if not name or name.lower() in existing_names:
            skipped += 1
            continue
        rtype = ROOM_TYPE_MAP.get(r["type"], RoomType.CLASSROOM)
        db.add(Room(
            room_id=_id(), college_id=college_id,
            name=name, capacity=r["capacity"],
            room_type=rtype,
            has_projector=r["projector"],
            has_computers=r["computers"],
            has_ac=r["ac"],
        ))
        existing_names.add(name.lower())
        inserted += 1

    await db.commit()
    return {"category": "rooms", "inserted": inserted, "skipped": skipped, "total": len(rows)}


async def _import_venues(db: AsyncSession, rows: list[dict], college_id: str):
    if not rows:
        return {"category": "venues", "inserted": 0, "skipped": 0, "total": 0}

    result = await db.execute(
        select(Venue.name).where(Venue.college_id == college_id)
    )
    existing_names = {n.lower() for n in result.scalars().all()}

    inserted, skipped = 0, 0

    for v in rows:
        name = v["name"].strip()
        if not name or name.lower() in existing_names:
            skipped += 1
            continue
        vtype = VENUE_TYPE_MAP.get(v["type"], VenueType.HALL)
        db.add(Venue(
            venue_id=_id(), college_id=college_id,
            name=name, capacity=v["capacity"],
            venue_type=vtype,
        ))
        existing_names.add(name.lower())
        inserted += 1

    await db.commit()
    return {"category": "venues", "inserted": inserted, "skipped": skipped, "total": len(rows)}


async def _import_timeslots(db: AsyncSession, rows: list[dict], college_id: str):
    if not rows:
        return {"category": "timeslots", "inserted": 0, "skipped": 0, "total": 0}

    # Check existing slot orders
    result = await db.execute(
        select(TimeSlotConfig.slot_order).where(TimeSlotConfig.college_id == college_id)
    )
    existing_orders = set(result.scalars().all())

    inserted, skipped = 0, 0

    for s in rows:
        if s["order"] in existing_orders:
            skipped += 1
            continue
        stype = SLOT_TYPE_MAP.get(s["type"], SlotType.LECTURE)
        db.add(TimeSlotConfig(
            slot_id=_id(), college_id=college_id,
            slot_order=s["order"], label=s["label"],
            start_time=s["start_time"], end_time=s["end_time"],
            slot_type=stype,
        ))
        existing_orders.add(s["order"])
        inserted += 1

    await db.commit()
    return {"category": "timeslots", "inserted": inserted, "skipped": skipped, "total": len(rows)}


async def _import_batches(db: AsyncSession, rows: list[dict], college_id: str):
    if not rows:
        return {"category": "batches", "inserted": 0, "skipped": 0, "total": 0}

    dept_map = await _get_dept_map(db, college_id)

    # Check existing batches
    result = await db.execute(
        select(Batch.dept_id, Batch.semester, Batch.name).join(
            Department, Batch.dept_id == Department.dept_id
        ).where(Department.college_id == college_id)
    )
    existing = {(row.dept_id, row.semester, row.name) for row in result.all()}

    inserted, skipped = 0, 0
    warnings = []

    for b in rows:
        dept_code = b["department"].strip().upper()
        if dept_code not in dept_map:
            warnings.append(f"Skipped batch {b['name']}: department '{dept_code}' not found")
            skipped += 1
            continue

        dept_id = dept_map[dept_code]
        key = (dept_id, b["semester"], b["name"])
        if key in existing:
            skipped += 1
            continue

        db.add(Batch(
            batch_id=_id(), dept_id=dept_id,
            semester=b["semester"], name=b["name"], size=b["size"],
        ))
        existing.add(key)
        inserted += 1

    await db.commit()
    return {"category": "batches", "inserted": inserted, "skipped": skipped, "total": len(rows), "warnings": warnings}
